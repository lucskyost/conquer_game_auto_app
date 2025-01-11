from ui.main_window_ui import Ui_MainWindow
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QApplication, QMainWindow, QGraphicsDropShadowEffect, QTableWidgetItem,QHBoxLayout, QTableWidget, QWidget, QHeaderView, QCheckBox, QMenu, QFileDialog, QMessageBox
from PySide6.QtCore import Qt
from openpyxl import load_workbook
from src.models.account import Account
class TableWidgetHandle:
    def __init__(self, mainWindowHandle):
        super().__init__()
        # self.ui = Ui_MainWindow()
        # self=mainWindowHandle
        self.mainWindowHandle= mainWindowHandle
        self.ui=mainWindowHandle.ui
        # self.ui = Ui_MainWindow()
        self.accounts = mainWindowHandle.accounts
        self.data_path=mainWindowHandle.data_path
        self.load_data_to_tableWidget(mainWindowHandle.data_path)
        self.ui.tableWidget.clicked.connect(self.on_tableWidget_clicked)
        self.ui.sellectAllCheckBox.stateChanged.connect(self.select_all_checkbox_handle)
    def load_data_to_tableWidget(self, filename):
        try:
            workbook = load_workbook(filename)
            sheet = workbook.active
            num_rows = sheet.max_row
            num_cols = sheet.max_column
            self.ui.tableWidget.setRowCount(num_rows)
            self.ui.tableWidget.setColumnCount(num_cols+2)
            for i in range(1, num_rows + 1):
                for j in range(1, num_cols + 1):
                    cell_value = sheet.cell(row=i, column=j).value
                    item = QTableWidgetItem(str(cell_value))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.ui.tableWidget.setItem(i-1 , j , item)
            self.set_status_collum()
            self.set_checkbox_collum()
            for column in range(self.ui.tableWidget.columnCount()):
                self.ui.tableWidget.resizeColumnToContents(column)
                self.ui.tableWidget.setColumnWidth(column, self.ui.tableWidget.columnWidth(column)+25) 
            self.ui.tableWidget.resizeColumnToContents(0)
            self.print_account_from_tableWidget(self.data_path)
        except FileNotFoundError:
            print(f"File '{filename}' not found.")
    def print_account_from_tableWidget(self, filename):
            try:
                workbook = load_workbook(filename)
                sheet = workbook.active
                num_rows = sheet.max_row
                self.accounts=[]
                for i in range(1, num_rows + 1):  # assuming first row is header
                    username = sheet.cell(row=i, column=1).value
                    password = sheet.cell(row=i, column=2).value
                    group = sheet.cell(row=i, column=3).value
                    server = sheet.cell(row=i, column=4).value
                    level = sheet.cell(row=i, column=5).value
                    gold = sheet.cell(row=i, column=6).value
                    status="inactive"
                    # status = sheet.cell(row=i, column=7).value
                    # # Convert status to boolean if needed
                    # if status.lower() == "active":
                    #     status = True
                    # else:
                    #     status = False
                    account = Account(username, password, group, server, level, gold, status)
                    self.accounts.append(account)
            except FileNotFoundError:
                print(f"File '{filename}' not found.")
                return None
            for account in self.accounts:
                print(account)
    def handle_checkbox_state_changed(self, checkbox):
        row = self.ui.tableWidget.indexAt(checkbox.parent().pos()).row()
        if checkbox.isChecked():
            print("Checkbox checked at row:", row)
        else:
            print("Checkbox unchecked at row:", row)
    def set_checkbox_collum(self):
        for row in range(self.ui.tableWidget.rowCount()):
            widget = QWidget()
            layout = QHBoxLayout(widget)
            layout.setContentsMargins(0, 0, 0, 0)
            checkbox = QCheckBox()
            layout.addWidget(checkbox)
            layout.setAlignment(checkbox, Qt.AlignCenter)
            widget.setLayout(layout)
            widget.setStyleSheet("""
                border-radius: 0;
                border: 1px;
            """)
            checkbox.setStyleSheet("""
                border-radius: 0;
                border: 1px;
            """)
            self.ui.tableWidget.setCellWidget(row, 0, widget)
            # checkbox.stateChanged.connect(self.handle_checkbox_state_changed)
            checkbox.stateChanged.connect(lambda state, checkbox=checkbox: self.handle_checkbox_state_changed(checkbox))
    def set_status_collum(self):
        column_index = self.ui.tableWidget.columnCount() - 1
        unique_value = "inactive"
        for row_index in range(self.ui.tableWidget.rowCount()):
            item = QTableWidgetItem(unique_value)
            self.ui.tableWidget.setItem(row_index, column_index, item)
    def on_tableWidget_clicked(self, item):
        # TableWidgetHandle(self).on_tableWidget_clicked(item) 
        cellContent = item.data()
        # print(cellContent)  # test
        sf = "You clicked on {0}x{1}".format(item.column(), item.row())
        usernameClicked=self.ui.tableWidget.item(item.row(), 1).text() 
        for account in self.accounts:
            if account.username==usernameClicked:
                # self.ui.autoFileLabel.setText("File sellected: " + account.selected_file)
                break
    def clear_highlight(self):
        # Để xóa highlight trên tất cả các dòng
        for row in range(self.ui.tableWidget.rowCount()):
            for column in range(self.ui.tableWidget.columnCount()):
                cell_widget = self.ui.tableWidget.item(row, column)
                if cell_widget:
                    cell_widget.setForeground(QColor("#dadbdf"))  # Đặt màu chữ về mặc định
    def hightLightRow(self, row, flag):
        # Bỏ qua nếu không có item được chọn
        # if item is None:
        #     return
        # # Lấy số dòng của item được chọn
        # row = item.row()
        # Deselect tất cả các hàng trước khi highlight hàng mới
        self.ui.tableWidget.clearSelection()
        for column in range(self.ui.tableWidget.columnCount()):
            cell_widget = self.ui.tableWidget.item(row, column)
            if cell_widget:
                if flag==True:
                    cell_widget.setForeground(Qt.yellow)  # Đặt màu chữ
                else:
                    cell_widget.setForeground(QColor("#dadbdf"))
    def select_all_checkbox_handle(self):
        if self.ui.sellectAllCheckBox.isChecked():
            for row in range(self.ui.tableWidget.rowCount()):
                item = self.ui.tableWidget.cellWidget(row, 0)
                # if isinstance(item, QWidget):
                checkbox = item.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(True)
        else:
            for row in range(self.ui.tableWidget.rowCount()):
                item = self.ui.tableWidget.cellWidget(row, 0)
                # if isinstance(item, QWidget):
                checkbox = item.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(False)